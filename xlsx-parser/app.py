from fastapi import FastAPI, UploadFile, File, Form
from openpyxl import load_workbook
import tempfile
import os

app = FastAPI()

@app.get("/health")
def health():
    return {"status": "ok"}

def get_font_color(cell):
    color = cell.font.color

    if color is None:
        return None

    if color.type == "rgb" and color.rgb:
        return color.rgb

    if color.type == "theme":
        return f"theme:{color.theme}"

    if color.type == "indexed":
        return f"indexed:{color.indexed}"

    return None

def get_fill_color(cell):
    fill = cell.fill

    if fill is None or fill.fgColor is None:
        return "00000000"

    if fill.fgColor.type == "rgb" and fill.fgColor.rgb:
        return fill.fgColor.rgb

    if fill.fgColor.type == "theme":
        return f"theme:{fill.fgColor.theme}"

    if fill.fgColor.type == "indexed":
        return f"indexed:{fill.fgColor.indexed}"

    return "00000000"

def read_sheet_with_styles(ws):
    rows = []

    for row in ws.iter_rows():
        row_items = []

        for cell in row:
            if cell.value is None:
                continue

            row_items.append({
                "cell": cell.coordinate,
                "value": cell.value,
                "bold": bool(cell.font.bold),
                "font_color": get_font_color(cell),
                "fill_color": get_fill_color(cell)
            })

        if row_items:
            rows.append(row_items)

    return rows

def parse_size_table(wb, target_product_code):
    if "사이즈표" not in wb.sheetnames:
        return {}, ""

    ws = wb["사이즈표"]
    rows = list(ws.iter_rows(values_only=True))

    result = {}
    current_code = None
    collecting = False

    for row in rows:
        row_values = list(row)

        first_cell = row_values[0] if len(row_values) > 0 else None

        if first_cell is not None and str(first_cell).strip() != "":
            current_code = str(first_cell).strip()
            collecting = current_code == target_product_code

        if not collecting:
            continue

        if len(row_values) < 3:
            continue

        label = row_values[1]

        if label is None or str(label).strip() == "":
            continue

        label_text = str(label).strip()

        values = []
        for value in row_values[2:]:
            if value is None or str(value).strip() == "":
                continue
            values.append(str(value).strip())

        if values:
            result[label_text] = values

    size_text = "\n".join(
        f"- {label}: {', '.join(values)}"
        for label, values in result.items()
    )

    return result, size_text

@app.post("/parse-xlsx")
async def parse_xlsx(
    excel: UploadFile = File(...),
    targetProductCode: str = Form("")
):
    suffix = os.path.splitext(excel.filename)[1]

    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await excel.read())
        tmp_path = tmp.name

    wb = load_workbook(tmp_path, data_only=True)

    sheet_names = wb.sheetnames

    if "정보표" in sheet_names:
        info_ws = wb["정보표"]
        selected_sheet_index = sheet_names.index("정보표")
    elif len(wb.worksheets) >= 3:
        info_ws = wb.worksheets[2]
        selected_sheet_index = 2
    else:
        info_ws = wb.worksheets[0]
        selected_sheet_index = 0

    info_rows = read_sheet_with_styles(info_ws)

    size_table, size_table_text = parse_size_table(wb, targetProductCode)

    os.remove(tmp_path)

    return {
        "sheet_names": sheet_names,
        "selected_sheet_index": selected_sheet_index,
        "sheet_name": info_ws.title,
        "targetProductCode": targetProductCode,
        "rows": info_rows,
        "sizeTable": size_table,
        "sizeTableText": size_table_text
    }